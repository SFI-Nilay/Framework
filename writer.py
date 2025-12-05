"""
writer.py
"""

import os
from openpyxl import Workbook, load_workbook
from typing import Dict

def _init_workbook(file_path: str) -> Workbook:
    """
    Initialize the Excel workbook if it does not exist, creating required sheets.
    """
    if os.path.exists(file_path):
        return load_workbook(file_path)

    wb = Workbook()

    # Sheet 1: Framework Overview
    ws1 = wb.active
    ws1.title = "Framework Overview"
    ws1.append([
        "Framework ID", "Issuer", "Framework Name", "SPO Provider", "Alignment",
        "Year", "SPO Date", "Framework Source"
    ])

    # Sheet 4: Governance
    ws4 = wb.create_sheet("Governance")
    ws4.append([
        "Framework ID", "Exclusion Criteria", "Impact Reporting", "External Verification"
    ])

    # Sheet 5: SPO Summary
    ws5 = wb.create_sheet("SPO Summary")
    ws5.append([
        "Framework ID", "Summary"
    ])

    wb.save(file_path)
    return wb


def _get_next_framework_id(ws) -> str:
    max_row = ws.max_row
    if max_row < 2:  # Only header present
        return "F001"
    last_id = ws.cell(row=max_row, column=1).value
    if not last_id or not str(last_id).startswith("F"):
        return "F001"
    number = int(last_id[1:])
    return f"F{number + 1:03d}"


def write_to_excel(json_data: Dict, run_for: str, file_path: str) -> None:
    """
    Write extracted JSON data into the Excel workbook at file_path.
    """
    wb = _init_workbook(file_path)
    ws1 = wb["Framework Overview"]

    # -------------------------
    # Determine Framework ID
    # -------------------------
    if run_for == "framework":
        # New framework, generate next ID
        framework_id = _get_next_framework_id(ws1)
    else:
        # SPO: attach to the last framework entered
        if ws1.max_row < 2:
            # If no framework exists, we can't attach an SPO.
            # You might want to handle this gracefully or log it.
            print("Warning: No framework entry exists yet. SPO data might be orphaned.")
            framework_id = "UNKNOWN"
        else:
            framework_id = ws1.cell(row=ws1.max_row, column=1).value

    # -------------------------
    # Write Framework Data
    # -------------------------
    if run_for == "framework":
        # Framework Overview
        ws1.append([
            framework_id,
            json_data.get("Issuer", ""),
            json_data.get("Framework Name", ""),
            json_data.get("SPO Provider", ""),
            json_data.get("Alignment", ""),
            json_data.get("Year", ""),
            json_data.get("SPO Date", ""),
            json_data.get("Framework Source", "")
        ])

        # Governance
        ws4 = wb["Governance"]
        ws4.append([
            framework_id,
            json_data.get("Exclusion Criteria", ""),
            json_data.get("Impact Reporting", ""),
            json_data.get("External Verification", "")
        ])

    # -------------------------
    # Write SPO Data
    # -------------------------
    elif run_for == "spo":
        # Update the last framework row in Framework Overview
        # (Only if we found a valid framework_id)
        if ws1.max_row >= 2:
            last_row = ws1.max_row
            ws1.cell(row=last_row, column=4, value=json_data.get("SPO Provider", ""))  # SPO Provider
            ws1.cell(row=last_row, column=7, value=json_data.get("SPO Date", ""))      # SPO Date

        # Append SPO Summary
        ws5 = wb["SPO Summary"]
        ws5.append([
            framework_id,
            json_data.get("Summary", "")
        ])

    wb.save(file_path)
    print(f"✅ Data written successfully for {run_for} ({framework_id}) to {file_path}")